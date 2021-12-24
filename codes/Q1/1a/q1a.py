import logging
import os
from os import path

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

class Q1a:

    combination_set=["PeriodKey","Category","Country","Subcategory","Segment"]

    def validate_expected_value(self,data,col_name,expected_values):
        df=data.copy()

        err_df=df[~df[col_name].isin(expected_values)]

        return err_df

    def rule_extra(self,data,col_expected_dict):

        df=data.copy()

        extra_list=[]
        all_df=None
        
        for key,val in col_expected_dict.items():
            err_df=self.validate_expected_value(df,key,val)          

            if len(err_df) >0:
                err_msg=f'Value at column {key} is not found in the list of expected values {val}'

                #err_df['Remarks'] = "err_msg"
                err_df.insert(0, 'Remarks', err_msg)
                extra_list.append(err_df)   
  
        if len(extra_list) > 0:
            all_err_df=pd.concat(extra_list)
            all_df=all_err_df.drop(columns=['Remarks'])

            all_df.drop_duplicates(keep="first",inplace=True)


        remarks_list=[]
        for index,row in all_df.iterrows(): 

            df_err=all_err_df
            for key in col_expected_dict:
                df_err=df_err[df_err[key] == row[key]]


            all_err = ','.join(df_err["Remarks"])

            remarks_list.append(all_err)


        if all_df is not None:
            all_df.insert(0, 'Remarks', remarks_list)


        return(all_df)      

    def rule_duplicate(self,data,unique_keys):

        df=data.copy()
   
        err_msg=f"duplicates for combination set {unique_keys}"
        err_df = df[df.duplicated(subset=unique_keys, keep='first')]
        if len(err_df) > 0:
            err_df.insert(0, 'Remarks', err_msg)
        

        return(err_df)

    def rule_sum(self,data,col_list,total):

        df=data.copy()

        df_new = df.copy()
        df_new["sum"]=0
        for col in col_list:
            df_new["sum"]=df_new["sum"] + df_new[col]

   

        err_df = df_new[df_new['sum'] != 100]

        err_msg=f"The sum of values for these columns {col_list} should be {total}% for each combination set."

        if len(err_df) > 0:
            err_df.insert(0, 'Remarks', err_msg)
        
        err_df=err_df.drop(columns=["sum"])

        return(err_df)

    def rule_range(self,data,col_list,min=None,max=None,inclusive=True):

        
        df=data.copy()

        err_df_list=[]
        for col in col_list:
            if min is not None and max is not None:
                min_df = df[df[col] < min] if inclusive else df[df[col] <= min]
                max_df= df[df[col] > max] if inclusive else df[df[col] >= max]
                err_df = pd.concat([min_df,max_df])
            elif min is not None:
                err_df = df[df[col] < min] if inclusive else df[df[col] <= min]
            elif max is not None:
                err_df = df[df[col] > max] if inclusive else df[df[col] >= max]
            else:
                err_df=None

            if err_df is not None:    
                err_df_list.append(err_df)

        all_err_df=None
        if len(err_df_list)>0:
            all_err_df=pd.concat(err_df_list)


            err_msg=f"The range of values for each of these columns {col_list} should be "
            if min is not None and max is not None:
                err_msg=err_msg+f"between {min} and {max}, both inclusive."
            elif min is not None:
                err_msg=err_msg+f"more than {min}, inclusive."
            else:
                err_msg=err_msg+f"less than {max}, inclusive."

            all_err_df.insert(0, 'Remarks', err_msg)

        return(all_err_df)

    def rule_logic_compare(self,data,col1,col2,compare_type="GT"):

        df=data.copy()

        err_df=None
        err_df = df[df[col1] <  df[col2]] if compare_type=="GT" else err_df
        err_df = df[df[col1] >  df[col2]] if compare_type=="LT" else err_df
        err_df = df[df[col1] != df[col2]] if compare_type=="EQ" else err_df


        err_msg=""
        err_msg=f"{col1} value should always be higher than {col2} value" if compare_type=="GT" else err_msg
        err_msg=f"{col1} value should always be less than {col2} value" if compare_type=="LT" else err_msg
        err_msg=f"{col1} value should always be equal to {col2} value" if compare_type=="EQ" else err_msg

        if len(err_df) > 0:
            err_df.insert(0, 'Remarks', err_msg)
        


        return(err_df)

    def rule_logic_total(self,data,total_col,total_col_val,sub_col,compare_type="GT"):
        df=data.copy()

        total_df=df[df[total_col] == total_col_val] 

        no_total_df=df[df[total_col] !=total_col_val] 

        match_cols=self.combination_set.copy()
        match_cols.remove(total_col)

        cols_compare = ["conversations","people","% positive","% negative","% neutral"]
        merge_df=pd.merge(total_df, no_total_df, on=match_cols)

        rename_dict={
        f'{total_col}_x':f'{total_col}',
        'conversations_x':'conversations', 
        'people_x': 'people',
        '% positive_x': '% positive',
        '% negative_x': '% negative',
        '% neutral_x': '% neutral'
        }
        merge_df.rename(columns=rename_dict, inplace = True)              

        err_df_list=[]
        for index,col in enumerate(cols_compare):
            err_df=merge_df[merge_df[f'{col}']<=merge_df[f'{col}_y']]#????
            err_msg=f"{total_col}='{total_col_val}' {col} value is not higher than '{col}' value for the same combination set."
            err_df.insert(0, 'Remarks', err_msg)

            if len(err_df)>0:
                err_df_list.append(err_df)

        all_err_df=None
        if len(err_df_list)>0:
            all_err_df=pd.concat(err_df_list)

        return(all_err_df)

    def rule_missing(self,data,col_name):
        #logging.debug(f"rule_missing---> col_name= {col_name}")
        df=data.copy()

        segment_list=["All", "16-23 years old", "24-39 years old", "40+ years old"]
        segment_list.sort()

        err_msg=f"There should always be four different segments available for each Subcategory, namely [Segment=’All’, ’16-23 years old’, ’24-39 years old’, ’40+ years old’]"
        
        unique_subcategories=df.Subcategory.unique()
 
        err_df_list=[]
        for val in unique_subcategories:
            subcategory_df=df[df["Subcategory"]==val]
            unique_segments=subcategory_df.Segment.unique()
            unique_segments.sort()

            diff=list(set(unique_segments) - set(segment_list))

            if  len(set(unique_segments)) < len(set(segment_list)) or len(diff)>0:
                err_df_list.append(subcategory_df)

        all_err_df=None
        if len(err_df_list)>0:
            all_err_df=pd.concat(err_df_list)
            all_err_df.insert(0, 'Remarks', err_msg)

        return(all_err_df)

    def output_excel(self,file_name, sheet_name, df,no_output_msg="no errors are found"):
  
        try:
            workbook = load_workbook(filename=file_name)         
        except FileNotFoundError:
            workbook = Workbook()

        sheet_list=workbook.sheetnames

        if sheet_name in sheet_list:
            sheet = workbook[sheet_name]
        else:
            sheet=workbook.create_sheet(sheet_name)


        if df is not None and len(df)>0:
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)
        else:        
            sheet["A1"] = no_output_msg
             
        if "Sheet" in workbook.sheetnames:
            del_sheet = workbook['Sheet']
            workbook.remove(del_sheet)            

        workbook.save(filename=file_name)        


    #===========================================================================================
    # Load Data from csv files from local
    #===========================================================================================
    def load_data(self,dir,file_type="csv",move_file=False,skip_rows=0,skip_footer=0):
        #logging.debug(f'load_data---> dir={dir}')

        err_msg=None
        #get from local csv files
        df_dict={}
  
        with os.scandir(dir) as fileList:
            for file in fileList:
                full_path=f'{dir}\{file.name}'
                logging.debug(f'---> full_path={full_path}')
                #logging.debug(f'---> path.isfile(full_path)={path.isfile(full_path)}') 
                #logging.debug(f'---> os.path.splitext(full_path)[1].lower()={os.path.splitext(full_path)[1].lower()==".csv"}')                               
                #only load csv file
                if (path.isfile(full_path) and os.path.splitext(full_path)[1].lower()==f'.{file_type}'):       
                    df = pd.read_csv(full_path, engine='python', skiprows=skip_rows, skipfooter=skip_footer)
                    #pd.read_csv(full_path, skiprows=17, skipfooter=1)

                    #clean data
                    df.fillna('', inplace=True) #e.g website can be blank

                    #store to list
                    df_dict[file.name]=df

        if df_dict:
            if move_file:
                #move processed files to processed directory
                self.move_files(self.CSV_DIR,self.PROCESSED_CSV_DIR)
                err_msg=None
        else:
            err_msg=f"No {file_type} files to process"

        return (df_dict,err_msg)

    #===========================================================================================
    # #V2 Move all processed files to processed folder in Local dir
    #============================================================================================    
    def move_files(self,from_dir,to_dir):
        #logging.debug(f"move_files ---> from_dir= {from_dir}")   
        
        with os.scandir(from_dir) as fileList:
            for file in fileList:
                full_path=f'{from_dir}\{file.name}'
                #only move csv file
                if (path.isfile(full_path) and os.path.splitext(full_path)[1].lower()==".csv"): 
                    logging.info(f'Moving {file.name}')

                    original_file=fr'{from_dir}\{file.name}'
                    new_file=fr'{to_dir}\{file.name}'

                    os.rename(original_file,new_file)

#main
log_format="[ %(asctime)s - %(levelname)s - %(threadName)s - (%(name)s - %(filename)s - %(funcName)s(), line %(lineno)d)]: %(message)s"
logging.basicConfig(filename='q1a.log', filemode='w', level=logging.DEBUG,format=log_format)

print(os.getcwd())
cwd=os.getcwd()

Q_DIR=rf"{cwd}\Q1"
CSV_DIR=rf"{Q_DIR}\1a"
PROCESSED_CSV_DIR=rf"{CSV_DIR}\processed"
OUTPUT_CSV_DIR=rf"{CSV_DIR}\output"

q1a=Q1a()

(df_list,err_msg)=q1a.load_data(CSV_DIR)#TEST
logging.info(f"parse ---> len(df_list): {len(df_list)}")

column_expected_dict={ 
    q1a.combination_set[0]:[1],
    q1a.combination_set[1]:["Technical Ability Test"],
    q1a.combination_set[2]:["Singapore"],
    q1a.combination_set[3]:["At home","In office","Total"],
    q1a.combination_set[4]:["All","16-23 years old","24-39 years old","40+ years old"]
}

filename="errors.xlsx"

for key, df in df_list.items():
    #logging.debug(f"---> key= {key}")
    #logging.debug(f"---> df= {df}")
 
    #####(1) dups
    rule_duplicate_errors=q1a.rule_duplicate(df,q1a.combination_set)
    logging.debug(f" ---> rule_duplicate_errors= {rule_duplicate_errors.to_string()}")  

    sheet_name="dups"
    q1a.output_excel(filename,sheet_name,rule_duplicate_errors)

    #####(2) missing
    output_list=[]
    rule_missing_errors=q1a.rule_missing(df,"Subcategory")
    logging.debug(f" ---> rule_missing_errors= {rule_missing_errors}") 

    output_list.append(rule_missing_errors) 

    #concat
    all_rule_missing_errors=pd.concat(output_list, axis=0)
    
    sheet_name="missing"
    q1a.output_excel(filename,sheet_name,all_rule_missing_errors) 

    #####(3) extra
    rule_extra_errors=q1a.rule_extra(df,column_expected_dict)
    logging.debug(f" ---> rule_extra_errors= {rule_extra_errors.to_string()}")  
 
    sheet_name="extra"
    q1a.output_excel(filename,sheet_name,rule_extra_errors)    

    ######(4) logic - 3

    output_list=[]

    #logic1
    rule_total2_errors=q1a.rule_logic_total(df,"Segment","All","Subcategory",compare_type="GT")
    logging.debug(f" ---> rule_total2_errors= {rule_total2_errors}") 
    output_list.append(rule_total2_errors) 
 
    #logic2
    rule_total1_errors=q1a.rule_logic_total(df,"Subcategory","Total","Segment",compare_type="GT")
    logging.debug(f" ---> rule_total1_errors= {rule_total1_errors.to_string()}") 
    output_list.append(rule_total1_errors) 

    #logic 3
    col_list=["conversations", "people"]
    rule_logic1_errors=q1a.rule_logic_compare(df,col_list[0],col_list[1],compare_type="GT")
    logging.debug(f" ---> rule_logic1_errors= {rule_logic1_errors.to_string()}")  
    output_list.append(rule_logic1_errors) 

    #concat
    all_rule_logic_errors=pd.concat(output_list, axis=0)

    sheet_name="logic"
    q1a.output_excel(filename,sheet_name,all_rule_logic_errors)  

    #####(5) sum
    col_list=["% positive", "% negative", "% neutral"]
    rule_sum_errors=q1a.rule_sum(df,col_list,100)
    logging.debug(f" ---> rule_sum_errors= {rule_sum_errors.to_string()}") 

    sheet_name="sum"
    q1a.output_excel(filename,sheet_name,rule_sum_errors)

    ######((6) range -2
    output_list=[]

    #range 1
    col_list=["% positive", "% negative", "% neutral"]
    rule_range1_errors=q1a.rule_range(df,col_list,min=0,max=100)
    logging.debug(f" ---> rule_range1_errors= {rule_range1_errors.to_string()}")
    output_list.append(rule_range1_errors) 

    #range 2
    col_list=["conversations", "people"]
    rule_range2_errors=q1a.rule_range(df,col_list,min=0)
    logging.debug(f" ---> rule_range2_errors= {rule_range2_errors.to_string()}") 
    output_list.append(rule_range2_errors) 

    #concat
    all_rule_range_errors=pd.concat(output_list, axis=0)

    sheet_name="range"
    q1a.output_excel(filename,sheet_name,all_rule_range_errors)
    
    
    #red_background = PatternFill(fgColor="00FF0000")
    #diff_style = DifferentialStyle(fill=red_background)
    #rule = Rule(type="expression", dxf=diff_style)
    #rule.formula = ["$H1<3"]
    #sheet.conditional_formatting.add("A1:O100", rule)

